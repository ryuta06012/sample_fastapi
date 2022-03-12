import shutil
import openpyxl
import os
from pathlib import Path

class InformationInput():
	def __init__(self, path : str):
		self.lsts = []
		self.expath = path
		self.name = 0
		self.column = 1
		self.row = 2

	def iterCols(self):
		cppath = shutil.copy(self.expath, './tmp')
		wb = openpyxl.load_workbook(cppath)
		sheet = wb[wb.sheetnames[0]]
		for	cols in sheet.iter_cols():
			for	cell in cols:
				if cell.value == '単価':
					self.lsts.append(['単価', cols[0].column, cell.row])
				elif cell.value == '数量':
					self.lsts.append(['数量', cols[0].column, cell.row])
				elif cell.value == '金額':
					self.lsts.append(['金額', cols[0].column, cell.row])
				elif cell.value == '合計金額':
					self.lsts.append(['合計金額', cols[0].column, cell.row])

	def addDataByColumn(self):
		cppath = shutil.copy(self.expath, './tmp')
		wb = openpyxl.load_workbook(cppath)
		sheet = wb[wb.sheetnames[0]]
		for lst in self.lsts:
			for	cols in sheet.iter_cols(min_col = lst[1], max_col=lst[1], min_row=lst[2],values_only=True):
				lst.append(cols)
		os.remove(cppath)
	
	def createExcel(self):
		filename = os.getcwd() + '/excel2/' + Path(self.expath).stem + '.xlsx'
		self.cxpath = shutil.copy('./format.xlsx', filename)
		wb = openpyxl.load_workbook(self.cxpath)
		sheet = wb[wb.sheetnames[0]]
		self.getFormatInfo(sheet)
		for lst in self.lsts:
			size = len(lst[3])
			if lst[self.name] == '単価':
				for i in range(size - 1):
					sheet.cell(column=self.price[self.column], row=self.price[self.row] + 1 + i).value = int(lst[3][i + 1])
			if lst[self.name] == '数量':
				for i in range(size - 1):
					sheet.cell(column=self.volume[self.column], row=self.volume[self.row] + 1 + i).value = int(lst[3][i + 1])
		wb.save(self.cxpath)
	
	def getPathName(self):
		basename = os.path.basename(self.cxpath)
		self.pathName_tuple = (basename, self.cxpath)
		return self.pathName_tuple

	def getFormatInfo(self, sheet):
		self.flsts = []
		for	cols in sheet.iter_cols():
			for	cell in cols:
				if cell.value == '単価':
					self.flsts.append(['単価', cols[0].column, cell.row])
				elif cell.value == '数量':
					self.flsts.append(['数量', cols[0].column, cell.row])
				elif cell.value == '金額':
					self.flsts.append(['金額', cols[0].column, cell.row])
				elif cell.value == '合計金額':
					self.flsts.append(['合計金額', cols[0].column, cell.row])
		self.divFormatInfo(self.flsts)

	def divFormatInfo(self, flsts : list):
		for flst in flsts:
			if flst[self.name] == '単価':
				self.price = flst
			if flst[self.name] == '数量':
				self.volume = flst
			if flst[self.name] == '金額':
				self.money = flst
			if flst[self.name] == '合計金額':
				self.total = flst